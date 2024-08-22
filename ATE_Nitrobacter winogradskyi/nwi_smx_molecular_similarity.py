from rdkit import Chem
from rdkit.Chem import AllChem, DataStructs, Draw
import os
import pandas as pd
from PIL import Image
from io import BytesIO
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Border, Side, Font, Alignment


def read_molecule(molecule_file_path):
    """Reads a molecule from a MOL file."""
    mol = Chem.MolFromMolFile(molecule_file_path)
    if mol is None:
        raise ValueError(f"Could not read molecule from file: {molecule_file_path}")
    return mol


def calculate_similarity(mol1, mol2):
    """Calculates the Tanimoto similarity between two molecules."""
    fp1 = AllChem.GetMorganFingerprintAsBitVect(mol1, 2)
    fp2 = AllChem.GetMorganFingerprintAsBitVect(mol2, 2)
    tanimoto_similarity = DataStructs.TanimotoSimilarity(fp1, fp2)
    return tanimoto_similarity


def draw_molecule(mol, highlight_atoms=None, highlight_bonds=None):
    """Draws a molecule with highlighted atoms and bonds."""
    if highlight_atoms is None:
        highlight_atoms = []
    if highlight_bonds is None:
        highlight_bonds = []

    # Draw the molecule
    drawer = Draw.MolDraw2DCairo(300, 300)
    drawer.DrawMolecule(mol, highlightAtoms=highlight_atoms, highlightBonds=highlight_bonds)
    drawer.FinishDrawing()
    img_data = drawer.GetDrawingText()
    return img_data


# Get the current directory
current_dir = os.path.dirname(os.path.abspath(__file__))

# Path to Sulfamethoxazole.mol
Sulfamethoxazole_file_path = os.path.join(current_dir, 'Sulfamethoxazole.mol')

# Read Sulfamethoxazole molecule
Sulfamethoxazole_molecule = read_molecule(Sulfamethoxazole_file_path)

# List to store results
results = []

# Iterate over all MOL files in the current directory
for mol_filename in os.listdir(current_dir):
    if mol_filename.endswith('.mol') and mol_filename != 'Sulfamethoxazole.mol':
        molecule_file_path = os.path.join(current_dir, mol_filename)
        try:
            mol = read_molecule(molecule_file_path)
            similarity_score = calculate_similarity(Sulfamethoxazole_molecule, mol)
            img_data = draw_molecule(mol)
            results.append((mol_filename, similarity_score, img_data))
        except Exception as e:
            print(f"Error processing {molecule_file_path}: {e}")

# Create a DataFrame from the results
df = pd.DataFrame(results, columns=['Molecule', 'Tanimoto Similarity', 'Image'])

# Save results to an Excel file
output_file_path = os.path.join(current_dir, 'nwi_smx_similarity_results.xlsx')

# Create a Pandas Excel writer using openpyxl as the engine
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    df[['Molecule', 'Tanimoto Similarity']].to_excel(writer, index=False, sheet_name='Results')
    workbook = writer.book
    worksheet = writer.sheets['Results']

    # Adjust column width
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 45  # Adjust column width for images

    # Add borders to all cells
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Add headline above the images column
    headline = "Molecular Similarity Results"
    worksheet.insert_rows(1)
    worksheet.cell(row=1, column=3).value = headline
    worksheet.cell(row=1, column=3).font = Font(size=14, bold=True)
    worksheet.cell(row=1, column=3).alignment = Alignment(horizontal='center')
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=3)

    # Add images and adjust row height
    for idx, (_, _, image_data) in enumerate(results, start=2):
        img = Image.open(BytesIO(image_data))
        openpyxl_img = OpenpyxlImage(img)
        openpyxl_img.anchor = f'C{idx + 1}'  # Place image starting from row 3 to leave space for the headline
        worksheet.add_image(openpyxl_img)
        worksheet.row_dimensions[idx + 1].height = img.height * 0.75  # Adjust row height based on image height

        # Add border to the cell with the image
        img_cell = worksheet.cell(row=idx + 1, column=3)
        img_cell.border = thin_border

print(f'Results saved to {output_file_path}')

# Create another Excel file with descending order of Tanimoto similarity
sorted_df = df.sort_values(by='Tanimoto Similarity', ascending=False).reset_index(drop=True)

output_sorted_file_path = os.path.join(current_dir, 'nwi_smx_similarity_results_sorted.xlsx')

with pd.ExcelWriter(output_sorted_file_path, engine='openpyxl') as writer:
    sorted_df[['Molecule', 'Tanimoto Similarity']].to_excel(writer, index=False, sheet_name='Results')
    workbook = writer.book
    worksheet = writer.sheets['Results']

    # Adjust column width
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 45  # Adjust column width for images

    # Add borders to all cells
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Add headline above the images column
    headline = "Molecular Similarity Results (Sorted)"
    worksheet.insert_rows(1)
    worksheet.cell(row=1, column=3).value = headline
    worksheet.cell(row=1, column=3).font = Font(size=14, bold=True)
    worksheet.cell(row=1, column=3).alignment = Alignment(horizontal='center')
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=3)

    # Add images and adjust row height
    for idx, (_, _, image_data) in enumerate(sorted_df.itertuples(index=False), start=2):
        img = Image.open(BytesIO(image_data))
        openpyxl_img = OpenpyxlImage(img)
        openpyxl_img.anchor = f'C{idx + 1}'  # Place image starting from row 3 to leave space for the headline
        worksheet.add_image(openpyxl_img)
        worksheet.row_dimensions[idx + 1].height = img.height * 0.75  # Adjust row height based on image height

        # Add border to the cell with the image
        img_cell = worksheet.cell(row=idx + 1, column=3)
        img_cell.border = thin_border

print(f'Sorted results saved to {output_sorted_file_path}')
